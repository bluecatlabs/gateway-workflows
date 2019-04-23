# Copyright 2019 BlueCat Networks (USA) Inc. and its affiliates
# -*- coding: utf-8 -*-
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# By: BlueCat Networks
# Date: 2019-03-14
# Gateway Version: 18.10.2
# Description: Zone Exporter exporter 

# Various Flask framework items.
import json
import csv
import ipaddress
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.utils.cell import column_index_from_string, get_column_letter


from bluecat import util
from bluecat.entity import Entity
from bluecat.api_exception import BAMException, PortalException

import config.default_config as config

col_config = {}
CONFIG_FILE = '%s/config_%s.json'

def module_path():
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_text():
    return util.get_text(module_path(), config.language)


r_text= get_resource_text()
SHEET_TITLE_FOR_CONFIGURATION = r_text['sheet_title_for_configuration']
SHEET_TITLE_FOR_VIEW = r_text['sheet_title_for_view']
SHEET_TITLE_FOR_ZONE = r_text['sheet_title_for_zone']
SHEET_TITLE_FOR_STRUCTURE = r_text['sheet_title_for_structure']

BOLD_FONT = Font(bold=True)
LINK_FONT = Font(underline='single', color=Color(theme=10, type="theme"))
BOLD_LINK_FONT = Font(bold=True, underline='single', color=Color(theme=10, type="theme"))
TITLE_FILL_COLOR = PatternFill('solid', fgColor='9BC2E6')

def load_config(module_dir, nodes):
    global col_config
    col_config = json.load(open(CONFIG_FILE % (module_dir, config.language)))
    for col in col_config['zone']:
        node = {}
        node['label'] = col['title']
        node['name'] = col['id']
        node['index'] = col['id']
        node['width'] = col['gwidth']
        node['sortable'] = False
        nodes.append(node)

def set_common_props(node, parent, level, entity):
    node['id'] = entity.get_id()
    node['parent'] = parent
    node['level'] = level + 1
    node['loaded'] = True

def set_props(node, entity):
    props = col_config['zone']
    for prop in props:
        id = prop['id']
        if id == 'name':
            node[id] = entity.get_name()
        else:
            node[id] = entity.get_property(id)

def construct_node(parent, level, entity):
    node = {}
    set_common_props(node, parent, level, entity)
    set_props(node, entity)
    node['isLeaf'] = False
    return node

def get_sorted_children(parent):
    nodes = []
    zones = []
    if parent.get_type() == Entity.Configuration:
        zones = parent.get_children_of_type(Entity.View)
    else:
        zones = parent.get_children_of_type(Entity.Zone)
    for zone in zones:
        node = {}
        node['entity'] = zone
        node['name'] = zone.get_name()
        nodes.append(node)
        
    nodes.sort(key = lambda node: node['name'])
    return nodes

def get_sorted_resource_records(zone):
    host_records = zone.get_children_of_type(Entity.HostRecord)
    resource_records = list(host_records)
    alias_records = zone.get_children_of_type(Entity.AliasRecord)
    resource_records.extend(list(alias_records))
    mx_records = zone.get_children_of_type(Entity.MXRecord)
    resource_records.extend(list(mx_records))
    txt_records = zone.get_children_of_type(Entity.TXTRecord)
    resource_records.extend(list(txt_records))
    srv_records = zone.get_children_of_type(Entity.SRVRecord)
    resource_records.extend(list(srv_records))
    generic_records = zone.get_children_of_type(Entity.GenericRecord)
    resource_records.extend(list(generic_records))
    hinfo_records = zone.get_children_of_type(Entity.HINFORecord)
    resource_records.extend(list(hinfo_records))
    naptr_records = zone.get_children_of_type(Entity.NAPTRRecord)
    resource_records.extend(list(naptr_records))
    
    resource_records.sort(key = lambda record: record.get_name())
    return resource_records
    
def get_record_data(resource_record):
    rd = ''
    if resource_record.get_type() == Entity.HostRecord:
        rd = resource_record.get_property('addresses')
    elif resource_record.get_type() == Entity.AliasRecord:
        rd = resource_record.get_linked_record_name()
    elif resource_record.get_type() == Entity.MXRecord:
        rd = '[%s]%s' % (resource_record.get_property('priority'),
                            resource_record.get_linked_record_name())
    elif resource_record.get_type() == Entity.TXTRecord:
        rd = resource_record.get_property('txt')
    elif resource_record.get_type() == Entity.HINFORecord:
        rd = '[%s][%s]' % (resource_record.get_property('cpu'),
                            resource_record.get_property('os'))
    elif resource_record.get_type() == Entity.SRVRecord:
        rd = '[%s][%s][%s]%s' % (resource_record.get_property('priority'),
                                    resource_record.get_property('weight'),
                                    resource_record.get_property('port'),
                                    resource_record.get_linked_record_name())
    elif resource_record.get_type() == Entity.NAPTRRecord:
        rd = '[%s][%s][%s][%s][%s][%s]' % (resource_record.get_property('order'),
                resource_record.get_property('preference'),
                resource_record.get_property('service'),
                resource_record.get_property('regexp'),
                resource_record.get_property('replacement'),
                resource_record.get_property('flags'))
    elif resource_record.get_type() == Entity.GenericRecord:
        rd = '[%s]%s' % (resource_record.get_property('type'),
                            resource_record.get_property('rdata'))
    return rd

def construct_structure_row(indent, entity):
    
    row = []
    props = col_config['zone']
    for prop in props:
        id = prop['id']
        if id == 'name':
            zone_name = ''
            for index in range(indent):
                zone_name += '\t'
            zone_name += entity.get_name()
            row.append(zone_name)
        else:
            row.append(entity.get_property(id))
    return row
    
# -----------------------------------
# Resource Record CSV writing functions.
# -----------------------------------
def write_header(writer):
    header = []
    props = col_config['resource_record']
    for prop in props:
        header.append(prop['title'])
    
    writer.writerow(header)
    
def write_resource_record(writer, resource_record):
    row = []
    props = col_config['resource_record']
    for prop in props:
        id = prop['id']
        if id == 'name':
            row.append(resource_record.get_name())
        elif id == 'type':
            row.append(resource_record.get_type())
        elif id == 'record_data':
            row.append(get_record_data(resource_record))
        else:
            row.append(resource_record.get_property(id))
    
    writer.writerow(row)
    
def write_resource_records(writer, resource_records):
    for resource_record in resource_records:
        write_resource_record(writer, resource_record)

def write_zone(writer, zone):
    resource_records = get_sorted_resource_records(zone)
    
    write_header(writer)
    write_resource_records(writer, resource_records)
    
def write_tree(api, writer, entity):
    if entity.get_type() == Entity.Zone:
        write_zone(writer, entity)
        
    nodes = get_sorted_children(entity)
    for node in nodes:
        write_tree(api, writer, node['entity'])
    
# --------------------------------------------
# Zone Structure CSV writing functions.
# --------------------------------------------
def write_structure_header(writer):
    header = []
    props = col_config['zone']
    for prop in props:
        header.append(prop['title'])
    
    writer.writerow(header)
    
def write_structure(api, writer, level, entity):
    nodes = get_sorted_children(entity)
    for node in nodes:
        row = construct_structure_row(level, node['entity'])
        writer.writerow(row)
        write_structure(api, writer, level + 1, node['entity'])

# --------------------------------------------
# Top Writing function for CSV format.
# --------------------------------------------
def export_as_csv(api, dirname, filename, id, contents):
    entity = api.get_entity_by_id(id)

    encoding = str(col_config['encoding'])
    f = open(dirname + '/' + filename, 'w', encoding=encoding)
    writer = csv.writer(f, lineterminator = '\n')
    
    if contents == 'struct' or contents == 'both':
        write_structure_header(writer)
        write_structure(api, writer, 0, entity)
        
    if contents == 'records' or contents == 'both':
        write_tree(api, writer, entity)
    
    f.close()

# -------------------------------------
# Resource Record Excel writing functions.
# -------------------------------------
def get_sheet_name(entity):
    name = entity.get_property('absoluteName')
    if name is None:
        name = entity.get_name()
    return name

def write_row_for_excel(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    for column in range(len(row)):
        sheet.cell(row=index, column=start_column_idx + column, value=row[column])
        
def set_fill_color(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    
    for column in range(len(row)):
        cell = sheet.cell(row=index, column=start_column_idx + column)
        cell.fill = TITLE_FILL_COLOR
        cell.font = BOLD_FONT
        
def set_column_width(sheet, start_column, widthes):
    start_column_idx = column_index_from_string(start_column)

    for index in range(len(widthes)):
        column_letter = get_column_letter(start_column_idx + index)
        sheet.column_dimensions[column_letter].width = widthes[index]

def add_auto_filters(sheet, start_column, start_row, length):
    start_column_idx = column_index_from_string(start_column)
    end_column  = get_column_letter(start_column_idx + length - 1)
    filter_range = start_column + str(start_row) +  ':' + end_column + str(start_row)
    sheet.auto_filter.ref = filter_range

def write_title_for_excel(sheet, entity, linked_cells):
    title_cell_index = col_config['excel']['title_cell_index']
    title_string = SHEET_TITLE_FOR_ZONE + get_sheet_name(entity)
    sheet[title_cell_index].value = title_string
    if (0 < len(linked_cells)) and (str(entity.get_id()) in linked_cells):
        sheet[title_cell_index].hyperlink = linked_cells[str(entity.get_id())]
        sheet[title_cell_index].font = BOLD_LINK_FONT
    else:
        sheet[title_cell_index].font = BOLD_FONT            
    
def write_header_for_excel(sheet, start_column, start_row):
    header = []
    widthes = []
    props = col_config['resource_record']
    for prop in props:
        header.append(prop['title'])
        widthes.append(prop['width'])
    
    write_row_for_excel(sheet, start_column, start_row, header)
    set_fill_color(sheet, start_column, start_row, header)
    set_column_width(sheet, start_column, widthes)
    sheet.freeze_panes = 'A' + str(start_row + 1)
    
def write_resource_record_for_excel(sheet, start_column, index, resource_record):
    row = []
    props = col_config['resource_record']
    for prop in props:
        id = prop['id']
        if id == 'name':
            row.append(resource_record.get_name())
        elif id == 'type':
            row.append(resource_record.get_type())
        elif id == 'record_data':
            row.append(get_record_data(resource_record))
        else:
            row.append(resource_record.get_property(id))
    write_row_for_excel(sheet, start_column, index, row)
    
def write_resource_records_for_excel(sheet, start_column, start_row, zone, resource_records):
    index = start_row + 1
    
    for resource_record in resource_records:
        write_resource_record_for_excel(sheet, start_column, index, resource_record)
        index += 1
        
def write_zone_for_excel(workbook, start_column, start_row, zone, linked_cells):
    resource_records = get_sorted_resource_records(zone)
    
    sheet = workbook.create_sheet(title=get_sheet_name(zone))
    write_title_for_excel(sheet, zone, linked_cells)
    write_header_for_excel(sheet, start_column, start_row)
    add_auto_filters(sheet, start_column, start_row, len(col_config['resource_record']))
    write_resource_records_for_excel(sheet, start_column, start_row, zone, resource_records)

def write_tree_for_excel(api, workbook, start_column, start_row, entity, linked_cells):
    if entity.get_type() == Entity.Zone:
        write_zone_for_excel(workbook, start_column, start_row, entity, linked_cells)
        
    nodes = get_sorted_children(entity)
    for node in nodes:
        write_tree_for_excel(api, workbook, start_column, start_row, node['entity'], linked_cells)
        
# --------------------------------------------
# Zone Structure Excel writing functions.
# --------------------------------------------
def write_structure_title_for_excel(sheet, entity, add_link, linked_cells):
    title_cell_index = col_config['excel']['title_cell_index']
    title_string = ''
    if entity.get_type() == Entity.Configuration:
        title_string = SHEET_TITLE_FOR_CONFIGURATION
    elif entity.get_type() == Entity.View:
        title_string = SHEET_TITLE_FOR_VIEW
    else:
        title_string = SHEET_TITLE_FOR_ZONE

    title_string += get_sheet_name(entity)

    sheet[title_cell_index].value = title_string
    
    if entity.get_type() == Entity.Zone and True == add_link:
        sheet[title_cell_index].hyperlink = "#'" + get_sheet_name(entity) + \
                    "'!" + col_config['excel']['title_cell_index']
        sheet[title_cell_index].font = LINK_FONT
        linked_cells[str(entity.get_id())] = "#'" + sheet.title + "'!" + title_cell_index
    else:
        sheet[title_cell_index].font = BOLD_FONT            

def write_structure_header_for_excel(start_column, start_row, sheet):
    header = []
    widthes = []
    props = col_config['zone']
    for prop in props:
        header.append(prop['title'])
        widthes.append(prop['width'])
    
    write_row_for_excel(sheet, start_column, start_row, header)
    set_fill_color(sheet, start_column, start_row, header)
    set_column_width(sheet, start_column, widthes)
    sheet.freeze_panes = 'A' + str(start_row + 1)

def write_structure_node_for_excel(api, sheet, level, start_column, index, entity, add_link, linked_cells):
    row = construct_structure_row(0, entity)
    write_row_for_excel(sheet, start_column, index, row)
    cell = sheet.cell(row=index, column=column_index_from_string(start_column))
    cell.alignment = Alignment(indent=level)
    if entity.get_type() == Entity.Zone and True == add_link:
        cell.hyperlink = "#'" + get_sheet_name(entity) + "'!" + col_config['excel']['title_cell_index']
        cell.font = LINK_FONT
        linked_cells[str(entity.get_id())] = "#'" + sheet.title + "'!" + start_column + str(index)
    
def write_structure_for_excel(api, sheet, level, start_column, index, entity, add_link, linked_cells):
    nodes = get_sorted_children(entity)
    for node in nodes:
        write_structure_node_for_excel(api, sheet, level, \
                                        start_column, index, node['entity'], add_link, linked_cells)
        if 0 < level:
            sheet.row_dimensions[index].outline_level = level
        index += 1
        index = write_structure_for_excel(api, sheet, level + 1, \
                                            start_column, index, node['entity'], add_link, linked_cells)
    return index

# --------------------------------------------
# Top Writing function for Excel format.
# --------------------------------------------
def export_as_excel(api, dirname, filename, id, contents):
    excel_config = col_config['excel']
    start_column = excel_config['start_column']
    start_row = excel_config['start_row']
    linked_cells = {}

    workbook = openpyxl.load_workbook(dirname + '/templates/MeiryoUI.xlsx')
    sheet = workbook.active
    entity = api.get_entity_by_id(id)
    if contents == 'struct' or contents == 'both':
        add_link = (True if contents == 'both' else False)
        sheet.title = get_sheet_name(entity) + ' ' + SHEET_TITLE_FOR_STRUCTURE
        write_structure_title_for_excel(sheet, entity, add_link, linked_cells)
        write_structure_header_for_excel(start_column, start_row, sheet)
        add_auto_filters(sheet, start_column, start_row, len(col_config['zone']))
        write_structure_for_excel(api, sheet, 0, start_column, start_row + 1, \
                                  entity, add_link, linked_cells)
    else:
        workbook.remove(sheet)
    
    if contents == 'records' or contents == 'both':
        write_tree_for_excel(api, workbook, start_column, start_row, entity, linked_cells)
    workbook.save(dirname + '/' + filename)
    
