# Copyright 2019 BlueCat Networks (USA) Inc. and its affiliates
# -*- coding: utf-8 -*-
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# By: BlueCat Networks
# Date: 2019-03-14
# Gateway Version: 18.10.2
# Description: Network Exporter exporter

# Various Flask framework items.
import json
import csv
import ipaddress
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from bluecat import util
from bluecat.entity import Entity
from bluecat.api_exception import BAMException, PortalException

import config.default_config as config

col_config = {}
CONFIG_FILE = '%s/config_%s.json'

def module_path():
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_text():
    return util.get_text(module_path(), config.language)

r_text = get_resource_text()
SHEET_TITLE_FOR_CONFIGURATION = r_text['sheet_title_for_configuration']
SHEET_TITLE_FOR_BLOCK = r_text['sheet_title_for_block']
SHEET_TITLE_FOR_NETWORK = r_text['sheet_title_for_network']

NETWORKID_DISPLAY_NAME = r_text['network_id_display_name']
BROADCAST_DISPLAY_NAME = r_text['broadcast_display_name']

BOLD_FONT = Font(bold=True)
LINK_FONT = Font(underline='single', color=Color(theme=10, type="theme"))
BOLD_LINK_FONT = Font(bold=True, underline='single', color=Color(theme=10, type="theme"))
TITLE_FILL_COLOR = PatternFill('solid', fgColor='9BC2E6')

def load_config(module_dir, nodes):
    global col_config
    col_config = json.load(open(CONFIG_FILE % (module_dir, config.language)))
    for col in col_config['range']:
        node = {}
        node['label'] = col['title']
        node['name'] = col['id']
        node['index'] = col['id']
        node['width'] = col['gwidth']
        node['sortable'] = False
        nodes.append(node)
            
def get_order(entity):
    order = 0
    range = entity.get_property('CIDR')
    if range is None:
        order = util.ip42int(entity.get_property('start'))
    else:
        val = range.split('/')
        order = util.ip42int(val[0])
    return order
    
def get_range(entity):
    if entity.get_type() == Entity.Configuration:
        return entity.get_name()
    else:
        range = entity.get_property('CIDR')
        if range is None:
            range = entity.get_property('start') + '-' + entity.get_property('end')
        return range

def get_sorted_children(parent):
    nodes = []
    ip4_blocks = parent.get_children_of_type(Entity.IP4Block)
    for ip4_block in ip4_blocks:
        node = {}
        node['entity'] = ip4_block
        node['order'] = get_order(ip4_block)
        nodes.append(node)
        
    ip4_networks = parent.get_children_of_type(Entity.IP4Network)
    for ip4_network in ip4_networks:
        node = {}
        node['entity'] = ip4_network
        node['order'] = get_order(ip4_network)
        nodes.append(node)
        
    nodes.sort(key = lambda node: node['order'])
    return nodes

def set_common_props(node, parent, level, entity):
    node['id'] = entity.get_id()
    node['order'] = get_order(entity)
    node['parent'] = parent
    node['level'] = level + 1
    node['loaded'] = True

def set_props(node, entity):
    props = col_config['range']
    for prop in props:
        id = prop['id']
        if id == 'range':
            node[id] = get_range(entity)
        elif id == 'name':
            node[id] = entity.get_name()
        elif id == 'type':
            node[id] = entity.get_type()
        else:
            node[id] = entity.get_property(id)
        
def construct_ip4_block_node(parent, level, entity):
    node = {}
    set_common_props(node, parent, level, entity)
    set_props(node, entity)
    node['isLeaf'] = False
    return node
    
def construct_ip4_network_node(parent, level, entity):
    node = {}
    set_common_props(node, parent, level, entity)
    set_props(node, entity)
    node['isLeaf'] = True
    return node

def construct_structure_row(indent, entity):
    row = []
    props = col_config['range']
    for prop in props:
        if prop['id'] == 'range':
            nw_range = ''
            for index in range(indent):
                nw_range += '\t'
            nw_range += get_range(entity)
            row.append(nw_range)
        elif prop['id'] == 'name':
            row.append(entity.get_name())
        elif prop['id'] == 'type':
            row.append(entity.get_type())
        else:
            row.append(entity.get_property(prop['id']))
    return row
    
# -----------------------------------
# IP Address CSV writing functions.
# -----------------------------------
def write_header(writer):
    header = []
    props = col_config['address']
    for prop in props:
        header.append(prop['title'])
    
    writer.writerow(header)
    
def write_blunk_address(writer, address, name):
    row = []
    props = col_config['address']
    for prop in props:
        id = prop['id']
        if id == 'ip_address':
            row.append(util.int2ip4(int(address)))
        elif id == 'name':
            row.append(name)
        else:
            row.append(None)
    writer.writerow(row)

def write_ip4_address(writer, ip4_address):
    row = []
    props = col_config['address']
    for prop in props:
        id = prop['id']
        if id == 'ip_address':
            row.append(ip4_address.get_address())
        elif id == 'name':
            row.append(ip4_address.get_name())
        else:
            row.append(ip4_address.get_property(id))
    writer.writerow(row)
    
def write_ip4_addresses(writer, ip4_network, ip4_addresses):
    for ip4_address in ip4_addresses:
        write_ip4_address(writer, ip4_address)

def write_full_ip4_addresses(writer, ip4_network, ip4_addresses):
    iter_address = iter(ip4_addresses)
    ip4_address = None
    try:
        ip4_address = next(iter_address)
    except StopIteration:
        pass
        
    cidr = ip4_network.get_property('CIDR')
    network = ipaddress.IPv4Network(cidr)
    start_address = util.ip42int(str(network.network_address))
    end_address = util.ip42int(str(network.broadcast_address))
    
    mask = int(cidr.split('/')[1])
    if mask <= 30:
        write_blunk_address(writer, start_address, NETWORKID_DISPLAY_NAME)
        start_address += 1
        end_address -= 1
        
    for address in range(start_address, end_address + 1):
        if ip4_address is not None and address == util.ip42int(ip4_address.get_address()):
            write_ip4_address(writer, ip4_address)
            try:
                ip4_address = next(iter_address)
            except StopIteration:
                ip4_address = None
                pass
        else:
            write_blunk_address(writer, address, '')
            
    if mask <= 30:
        write_blunk_address(writer, end_address + 1, BROADCAST_DISPLAY_NAME)

def write_ip4_network(writer, ip4_network, full):
    children = ip4_network.get_children_of_type(Entity.IP4Address)
    ip4_addresses = list(children)
    ip4_addresses.sort(key = lambda e: util.ip42int(e.get_address()))
    
    write_header(writer)
    if full == True:
        write_full_ip4_addresses(writer, ip4_network, ip4_addresses)
    else:
        write_ip4_addresses(writer, ip4_network, ip4_addresses)
    
def write_tree(api, writer, entity, full):
    if entity.get_type() == Entity.IP4Network:
        write_ip4_network(writer, entity, full)
    else:
        nodes = get_sorted_children(entity)
        for node in nodes:
            write_tree(api, writer, node['entity'], full)
        
# --------------------------------------------
# Network Structure CSV writing functions.
# --------------------------------------------
def write_structure_header(writer):
    header = []
    props = col_config['range']
    for prop in props:
        header.append(prop['title'])
    
    writer.writerow(header)
    
def write_structure(api, writer, level, entity):
    if entity.get_type() != Entity.IP4Network:
        nodes = get_sorted_children(entity)
        if 0 < len(nodes):
            for node in nodes:
                row = construct_structure_row(level, node['entity'])
                writer.writerow(row)
                write_structure(api, writer, level + 1, node['entity'])

# --------------------------------------------
# Top Writing function for CSV format.
# --------------------------------------------
def export_as_csv(api, dirname, filename, id, contents, full):
    entity = api.get_entity_by_id(id)

    encoding = str(col_config['encoding'])
    f = open(dirname + '/' + filename, 'w', encoding=encoding)
    writer = csv.writer(f, lineterminator = '\n')
    
    if contents == 'struct' or contents == 'both':
        write_structure_header(writer)
        write_structure(api, writer, 0, entity)
        
    if contents == 'ip' or contents == 'both':
        write_tree(api, writer, entity, full)
    f.close()

# -------------------------------------
# IP Address Excel writing functions.
# -------------------------------------
def get_sheet_name(entity):
    return get_range(entity).replace('/', ' MASK ')

def write_row_for_excel(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    for column in range(len(row)):
        sheet.cell(row=index, column=start_column_idx + column, value=row[column])
        
def set_fill_color(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    
    for column in range(len(row)):
        cell = sheet.cell(row=index, column=start_column_idx + column)
        cell.fill = TITLE_FILL_COLOR
        cell.font = BOLD_FONT
        
def set_column_width(sheet, start_column, widthes):
    start_column_idx = column_index_from_string(start_column)

    for index in range(len(widthes)):
        column_letter = get_column_letter(start_column_idx + index)
        sheet.column_dimensions[column_letter].width = widthes[index]

def add_auto_filters(sheet, start_column, start_row, length):
    start_column_idx = column_index_from_string(start_column)
    end_column  = get_column_letter(start_column_idx + length - 1)
    filter_range = start_column + str(start_row) +  ':' + end_column + str(start_row)
    sheet.auto_filter.ref = filter_range

def write_title_for_excel(sheet, entity, linked_cells):
    title_cell_index = col_config['excel']['title_cell_index']
    title_string = SHEET_TITLE_FOR_NETWORK + get_range(entity)
    if entity.get_name() != None:
        title_string += ' - ' + entity.get_name()
    sheet[title_cell_index].value = title_string
    if 0 < len(linked_cells):
        sheet[title_cell_index].hyperlink = linked_cells[str(entity.get_id())]
        sheet[title_cell_index].font = BOLD_LINK_FONT
    else:
        sheet[title_cell_index].font = BOLD_FONT            
    
def write_header_for_excel(sheet, start_column, start_row):
    header = []
    widthes = []
    props = col_config['address']
    for prop in props:
        header.append(prop['title'])
        widthes.append(prop['width'])
    
    write_row_for_excel(sheet, start_column, start_row, header)
    set_fill_color(sheet, start_column, start_row, header)
    set_column_width(sheet, start_column, widthes)
    sheet.freeze_panes = 'A' + str(start_row + 1)
    
def write_blunk_address_for_excel(sheet, start_column, index, address, name):
    start_column_idx = column_index_from_string(start_column)

    sheet.cell(row=index, column=start_column_idx, value=util.int2ip4(int(address)))
    sheet.cell(row=index, column=start_column_idx + 1, value=name)
    
def write_ip4_address_for_excel(sheet, start_column, index, ip4_address):
    row = []
    props = col_config['address']
    for prop in props:
        id = prop['id']
        if id == 'ip_address':
            row.append(ip4_address.get_address())
        elif id == 'name':
            row.append(ip4_address.get_name())
        else:
            row.append(ip4_address.get_property(id))
            
    write_row_for_excel(sheet, start_column, index, row)
    
def write_ip4_addresses_for_excel(sheet, start_column, start_row, ip4_network, ip4_addresses):
    index = start_row + 1
    
    for ip4_address in ip4_addresses:
        write_ip4_address_for_excel(sheet, start_column, index, ip4_address)
        index += 1
        
def write_full_ip4_addresses_for_excel(sheet, start_column, start_row, ip4_network, ip4_addresses):
    index = start_row + 1
    iter_address = iter(ip4_addresses)
    ip4_address = None
    try:
        ip4_address = next(iter_address)
    except StopIteration:
        pass
        
    cidr = ip4_network.get_property('CIDR')
    network = ipaddress.IPv4Network(cidr)
    start_address = util.ip42int(str(network.network_address))
    end_address = util.ip42int(str(network.broadcast_address))
    
    mask = int(cidr.split('/')[1])
    if mask <= 30:
        write_blunk_address_for_excel(sheet, start_column, index, start_address, NETWORKID_DISPLAY_NAME)
        start_address += 1
        end_address -= 1
        index += 1
        
    for address in range(start_address, end_address + 1):
        if ip4_address is not None and address == util.ip42int(ip4_address.get_address()):
            write_ip4_address_for_excel(sheet, start_column, index, ip4_address)
            try:
                ip4_address = next(iter_address)
            except StopIteration:
                pass
        else:
            write_blunk_address_for_excel(sheet, start_column, index, address, '')
        index += 1
        
    if mask <= 30:
        write_blunk_address_for_excel(sheet, start_column, index, end_address + 1, BROADCAST_DISPLAY_NAME)

def write_ip4_network_for_excel(workbook, start_column, start_row, ip4_network, full, linked_cells):
    children = ip4_network.get_children_of_type(Entity.IP4Address)
    ip4_addresses = list(children)
    ip4_addresses.sort(key = lambda e: util.ip42int(e.get_address()))
    
    sheet = workbook.create_sheet(title=get_sheet_name(ip4_network))
    write_title_for_excel(sheet, ip4_network, linked_cells)
    write_header_for_excel(sheet, start_column, start_row)
    add_auto_filters(sheet, start_column, start_row, len(col_config['address']))

    if full == True:
        write_full_ip4_addresses_for_excel(sheet, start_column, start_row, ip4_network, ip4_addresses)
    else:
        write_ip4_addresses_for_excel(sheet, start_column, start_row, ip4_network, ip4_addresses)

def write_tree_for_excel(api, workbook, start_column, start_row, entity, full, linked_cells):
    if entity.get_type() == Entity.IP4Network:
        write_ip4_network_for_excel(workbook, start_column, start_row, entity, full, linked_cells)
    else:
        nodes = get_sorted_children(entity)
        for node in nodes:
            write_tree_for_excel(api, workbook, start_column, start_row, node['entity'], full, linked_cells)

# --------------------------------------------
# Network Structure Excel writing functions.
# --------------------------------------------
def write_structure_title_for_excel(sheet, entity):
    title_cell_index = col_config['excel']['title_cell_index']
    title_string = ''
    if entity.get_type() == Entity.Configuration:
        title_string = SHEET_TITLE_FOR_CONFIGURATION + entity.get_name()
    else:
        if entity.get_type() == Entity.IP4Block:
            title_string = SHEET_TITLE_FOR_BLOCK
        else:
            title_string = SHEET_TITLE_FOR_NETWORK
        title_string += get_range(entity)
        if entity.get_name() != None:
            title_string += ' - ' + entity.get_name()
            
    sheet[title_cell_index].value = title_string
    sheet[title_cell_index].font = BOLD_FONT            

def write_structure_header_for_excel(start_column, start_row, sheet):
    header = []
    widthes = []
    props = col_config['range']
    for prop in props:
        header.append(prop['title'])
        widthes.append(prop['width'])
    
    write_row_for_excel(sheet, start_column, start_row, header)
    set_fill_color(sheet, start_column, start_row, header)
    set_column_width(sheet, start_column, widthes)
    sheet.freeze_panes = 'A' + str(start_row + 1)

def write_structure_node_for_excel(api, sheet, level, start_column, index, entity, add_link, linked_cells):
    row = construct_structure_row(0, entity)
    write_row_for_excel(sheet, start_column, index, row)
    cell = sheet.cell(row=index, column=column_index_from_string(start_column))
    cell.alignment = Alignment(indent=level)
    if entity.get_type() == Entity.IP4Network and True == add_link:
        cell.hyperlink = "#'" + get_sheet_name(entity) + "'!" + col_config['excel']['title_cell_index']
        cell.font = LINK_FONT
        linked_cells[str(entity.get_id())] = "#'" + sheet.title + "'!" + start_column + str(index)
    
def write_structure_for_excel(api, sheet, level, start_column, index, entity, add_link, linked_cells):
    if entity.get_type() != Entity.IP4Network:
        nodes = get_sorted_children(entity)
        
        if 0 < len(nodes):
            for node in nodes:
                write_structure_node_for_excel(api, sheet, level, \
                                               start_column, index, node['entity'], add_link, linked_cells)
                if 0 < level:
                    sheet.row_dimensions[index].outline_level = level
                index += 1
                index = write_structure_for_excel(api, sheet, level + 1, \
                                                  start_column, index, node['entity'], add_link, linked_cells)
    return index

# --------------------------------------------
# Top Writing function for Excel format.
# --------------------------------------------
def export_as_excel(api, dirname, filename, id, contents, full):
    excel_config = col_config['excel']
    start_column = excel_config['start_column']
    start_row = excel_config['start_row']
    linked_cells = {}

    workbook = openpyxl.load_workbook(dirname + '/templates/MeiryoUI.xlsx')
    sheet = workbook.active
    entity = api.get_entity_by_id(id)
    if contents == 'struct' or contents == 'both':
        sheet.title = get_sheet_name(entity)
        write_structure_title_for_excel(sheet, entity)
        write_structure_header_for_excel(start_column, start_row, sheet)
        add_auto_filters(sheet, start_column, start_row, len(col_config['range']))
        write_structure_for_excel(api, sheet, 0, start_column, start_row + 1, \
                                  entity, (True if contents == 'both' else False), linked_cells)
    else:
        workbook.remove(sheet)
    
    if contents == 'ip' or contents == 'both':
        write_tree_for_excel(api, workbook, start_column, start_row, entity, full, linked_cells)
    workbook.save(dirname + '/' + filename)
    