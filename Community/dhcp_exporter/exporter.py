# Copyright 2021 BlueCat Networks (USA) Inc. and its affiliates
# -*- coding: utf-8 -*-
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# By: BlueCat Networks
# Date: 2021-05-01
# Gateway Version: 21.5.1
# Description: DHCP Exporter Page

# Various Flask framework items.
import json
import csv
import ipaddress
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from bluecat import util
from bluecat.entity import Entity
from bluecat.api_exception import BAMException, PortalException

from flask import g

import config.default_config as config

# Workaround for DHCPV4ClientOption, DHCPServiceOption
def get_deployment_options(entity, option_types, server_id):
    options = []
    try:
        options = entity._api_client.getDeploymentOptions(entity.get_id(), option_types, server_id)
    except GeneralError as e:
        print('Exception at get_deployment_options(%s)' % util.safe_str(e))
        raise BAMException(safe_str(e))
    return options
    
def get_value_from_properties(properties, key):
    props = re.split(r'(?<!\\)\|', properties)
    for prop in props:
        value = prop.split('=')
        if value[0] == key:
            return value[1]
    return ''
    

col_config = {}
CONFIG_FILE = '%s/config_%s.json'

def module_path():
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_text():
    return util.get_text(module_path(), config.language)

r_text = get_resource_text()
SHEET_TITLE_FOR_CONFIGURATION = r_text['sheet_title_for_configuration']
SHEET_TITLE_FOR_BLOCK = r_text['sheet_title_for_block']
SHEET_TITLE_FOR_NETWORK = r_text['sheet_title_for_network']

NETWORKID_DISPLAY_NAME = r_text['network_id_display_name']
BROADCAST_DISPLAY_NAME = r_text['broadcast_display_name']

BOLD_FONT = Font(bold=True)
LINK_FONT = Font(underline='single', color=Color(theme=10, type="theme"))
BOLD_LINK_FONT = Font(bold=True, underline='single', color=Color(theme=10, type="theme"))
TITLE_FILL_COLOR = PatternFill('solid', fgColor='9BC2E6')

def load_config(module_dir, nodes):
    global col_config
    col_config = json.load(open(CONFIG_FILE % (module_dir, config.language)))
    for col in col_config['props']:
        if col['gwidth'] == 0:
            continue
            
        node = {}
        node['label'] = col['title']
        node['name'] = col['id']
        node['index'] = col['id']
        node['width'] = col['gwidth']
        node['sortable'] = False
        nodes.append(node)
            
def get_order(entity):
    order = 0
    etype = entity.get_type()
    if etype in [Entity.IP4Block, Entity.IP4Network]:
        range = entity.get_property('CIDR')
        if range is not None:
            val = range.split('/')
            order = util.ip42int(val[0])
        else:
            order = util.ip42int(entity.get_property('start'))
    elif etype == Entity.DHCP4Range:
        order = util.ip42int(entity.get_property('start'))
    elif etype == Entity.IP4Address:
        order = util.ip42int(entity.get_address())
    return order
    
def get_range(entity):
    range = '-'
    etype = entity.get_type()
    if etype == Entity.Configuration:
        range = entity.get_name()
    elif etype in [Entity.IP4Block, Entity.IP4Network]:
        range = entity.get_property('CIDR')
        if range is None:
            range = entity.get_property('start') + '-' + entity.get_property('end')
    elif etype == Entity.DHCP4Range:
        range = entity.get_property('start') + '-' + entity.get_property('end')
    elif etype == Entity.IP4Address:
        range = entity.get_address()
    return range

def get_sorted_children(parent):
    nodes = []
    for ip4_block in parent.get_children_of_type(Entity.IP4Block):
        node = {}
        node['entity'] = ip4_block
        node['order'] = get_order(ip4_block)
        nodes.append(node)
        
    for ip4_network in parent.get_children_of_type(Entity.IP4Network):
        node = {}
        node['entity'] = ip4_network
        node['order'] = get_order(ip4_network)
        nodes.append(node)
        
    for dhcp4_range in parent.get_children_of_type(Entity.DHCP4Range):
        node = {}
        node['entity'] = dhcp4_range
        node['order'] = get_order(dhcp4_range)
        nodes.append(node)
        
    for ip4_address in parent.get_children_of_type(Entity.IP4Address):
        node = {}
        if ip4_address.get_state() == 'DHCP_RESERVED':
            node['entity'] = ip4_address
            node['order'] = get_order(ip4_address)
            nodes.append(node)
        
    nodes.sort(key = lambda node: node['order'])
    return nodes

def is_inherited(option):
    result = False
    inherited = get_value_from_properties(option['properties'], 'inherited')
    return True if inherited == 'true' else False
    
mac_pools = {}
dhcp_classes = {}

def clear_pool_class_cache():
    global mac_pools, dhcp_classes
    mac_pools = {}
    dhcp_classes = {}

def get_mac_pool_name(properties):
    api = api = g.user.get_api()
    global mac_pools
    id = get_value_from_properties(properties, 'macPool')
    
    pool_name = None
    if id in mac_pools.keys():
        pool_name = mac_pools[id]
    else:
        entity = api.get_entity_by_id(id)
        if entity is not None:
            pool_name = entity.get_name()
            mac_pools[id] = pool_name
    return pool_name

def get_dhcp_class_name(properties):
    api = api = g.user.get_api()
    global dhcp_classes
    id = get_value_from_properties(properties, 'dhcpMatchClass')
    
    class_name = None
    if id in dhcp_classes.keys():
        class_name = mac_pools[id]
    else:
        entity = api.get_entity_by_id(id)
        if entity is not None:
            class_name = entity.get_name()
            dhcp_classes[id] = class_name
    return class_name

def collect_options(entity):
    options = []
    try:
        for dop in get_deployment_options(entity, 'DHCPV4ClientOption', -1):
            if not is_inherited(dop):
                options.append(dop['name'] + '=' + dop['value'])
        for dop in get_deployment_options(entity, 'DHCPServiceOption', -1):
            if not is_inherited(dop):
                if '-pool' in dop['name']:
                    pool_name = get_mac_pool_name(dop['properties'])
                    options.append(dop['name'] + '=' + pool_name)
                elif '-dhcp-class-' in dop['name']:
                    print(dop)
                    class_name = get_dhcp_class_name(dop['properties'])
                    options.append(dop['name'] + '=' + class_name)
                else:
                    options.append(dop['name'] + '=' + dop['value'])
    except PortalException as pe:
        print(pe)
        
    return options

def set_common_props(node, parent, level, entity):
    node['id'] = entity.get_id()
    node['order'] = get_order(entity)
    node['parent'] = parent
    node['level'] = level + 1
    node['loaded'] = True

def set_props(node, entity, gui_flag = False):
    props = col_config['props']
    type_names = col_config['type_names']
    for prop in props:
        if gui_flag and prop['gwidth'] == 0:
            continue
            
        id = prop['id']
        if id == 'range':
            node[id] = get_range(entity)
        elif id == 'name':
            node[id] = entity.get_name()
        elif id == 'type':
            node[id] = type_names[entity.get_type()]
        else:
            node[id] = entity.get_property(id)
        
def construct_ip4_block_node(parent, level, entity):
    node = {}
    set_common_props(node, parent, level, entity)
    set_props(node, entity, True)
    node['isLeaf'] = False
    return node
    
def construct_ip4_network_node(parent, level, entity):
    node = {}
    set_common_props(node, parent, level, entity)
    set_props(node, entity, True)
    node['isLeaf'] = True
    return node

def construct_structure_row(indent, entity):
    row = []
    props = col_config['props']
    type_names = col_config['type_names']
    for prop in props:
        if prop['id'] == 'range':
            nw_range = ''
            for index in range(indent):
                nw_range += '\t'
            nw_range += get_range(entity)
            row.append(nw_range)
        elif prop['id'] == 'name':
            row.append(entity.get_name())
        elif prop['id'] == 'type':
            row.append(type_names[entity.get_type()])
        elif prop['id'] == 'options':
            row.append('\n'.join(collect_options(entity)))
        else:
            row.append(entity.get_property(prop['id']))
    return row
    
def is_in_dhcp_range(dhcp_ranges, address):
    result = False
    for dhcp_range in dhcp_ranges:
        start_ip = util.ip42int(dhcp_range.get_property('start'))
        end_ip = util.ip42int(dhcp_range.get_property('end'))
        ip4_address = util.int2ip4(address)
        if start_ip <= address and address <= end_ip:
            result = True
            break
    return result

def has_dhcp_range_or_reserve(entity):
    result = False
    if Entity.IP4Block == entity.get_type():
        for ip4_block in entity.get_ip4_blocks():
            result = has_dhcp_range_or_reserve(ip4_block)
            if result:
                break
                
        if result == False:
            for ip4_network in entity.get_ip4_networks():
                result = has_dhcp_range_or_reserve(ip4_network)
                if result:
                    break
                    
    elif Entity.IP4Network == entity.get_type():
        dhcp_ranges = entity.get_children_of_type(Entity.DHCP4Range)
        if 0 < len(list(dhcp_ranges)):
            result = True
        else:
            for ip4_address in entity.get_children_of_type(Entity.IP4Address):
                if ip4_address.get_state() == 'DHCP_RESERVED':
                    result = True
                    break
    return result

# --------------------------------------------
# Network Structure CSV writing functions.
# --------------------------------------------
def write_structure_header(writer):
    header = []
    props = col_config['props']
    for prop in props:
        header.append(prop['title'])
    
    writer.writerow(header)
    
def write_structure(writer, level, entity, full):
    etype = entity.get_type()
    if etype in [Entity.DHCP4Range, Entity.IP4Address]:
        row = construct_structure_row(level, entity)
        writer.writerow(row)
        
    elif has_dhcp_range_or_reserve(entity):
        if etype == Entity.IP4Network or full:
            row = construct_structure_row(level, entity)
            writer.writerow(row)
            level += 1
            
        nodes = get_sorted_children(entity)
        for node in nodes:
            write_structure(writer, level, node['entity'], full)

# --------------------------------------------
# Top Writing function for CSV format.
# --------------------------------------------
def export_as_csv(api, dirname, filename, id, full):
    clear_pool_class_cache()
    entity = api.get_entity_by_id(id)

    encoding = str(col_config['encoding'])
    f = open(dirname + '/' + filename, 'w', encoding=encoding)
    writer = csv.writer(f, lineterminator = '\n')
    
    write_structure_header(writer)
    if entity.get_type() == Entity.Configuration:
        for ip4_block in entity.get_children_of_type(Entity.IP4Block):
            write_structure(writer, 0, ip4_block, full)
    else:
        write_structure(writer, 0, entity, full)
        
    f.close()

# -------------------------------------
# IP Address Excel writing functions.
# -------------------------------------
def get_sheet_name(entity):
    return get_range(entity).replace('/', ' MASK ')

def write_row_for_excel(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    for column in range(len(row)):
        cell_value = row[column]
        cell = sheet.cell(row=index, column=start_column_idx + column, value=cell_value)
        if (cell_value is not None) and (1 < len(cell_value.split('\n'))):
            cell.alignment = Alignment(wrapText=True)
            
        
def set_fill_color(sheet, start_column, index, row):
    start_column_idx = column_index_from_string(start_column)
    
    for column in range(len(row)):
        cell = sheet.cell(row=index, column=start_column_idx + column)
        cell.fill = TITLE_FILL_COLOR
        cell.font = BOLD_FONT
        
def set_column_width(sheet, start_column, widthes):
    start_column_idx = column_index_from_string(start_column)

    for index in range(len(widthes)):
        column_letter = get_column_letter(start_column_idx + index)
        sheet.column_dimensions[column_letter].width = widthes[index]

def add_auto_filters(sheet, start_column, start_row, length):
    start_column_idx = column_index_from_string(start_column)
    end_column  = get_column_letter(start_column_idx + length - 1)
    filter_range = start_column + str(start_row) +  ':' + end_column + str(start_row)
    sheet.auto_filter.ref = filter_range

# --------------------------------------------
# Network Structure Excel writing functions.
# --------------------------------------------
def write_structure_title_for_excel(sheet, entity):
    title_cell_index = col_config['excel']['title_cell_index']
    title_string = ''
    if entity.get_type() == Entity.Configuration:
        title_string = SHEET_TITLE_FOR_CONFIGURATION + entity.get_name()
    else:
        if entity.get_type() == Entity.IP4Block:
            title_string = SHEET_TITLE_FOR_BLOCK
        else:
            title_string = SHEET_TITLE_FOR_NETWORK
        title_string += get_range(entity)
        if entity.get_name() != None:
            title_string += ' - ' + entity.get_name()
            
    sheet[title_cell_index].value = title_string
    sheet[title_cell_index].font = BOLD_FONT            

def write_structure_header_for_excel(start_column, start_row, sheet):
    header = []
    widthes = []
    props = col_config['props']
    for prop in props:
        header.append(prop['title'])
        widthes.append(prop['width'])
    
    write_row_for_excel(sheet, start_column, start_row, header)
    set_fill_color(sheet, start_column, start_row, header)
    set_column_width(sheet, start_column, widthes)
    sheet.freeze_panes = 'A' + str(start_row + 1)

def write_structure_node_for_excel(sheet, level, start_column, index, entity):
    row = construct_structure_row(0, entity)
    write_row_for_excel(sheet, start_column, index, row)
    cell = sheet.cell(row=index, column=column_index_from_string(start_column))
    cell.alignment = Alignment(indent=level, vertical='center')
    
def write_structure_for_excel(sheet, level, start_column, index, entity, full):
    etype = entity.get_type()
    if etype in [Entity.DHCP4Range, Entity.IP4Address]:
        write_structure_node_for_excel(sheet, level, start_column, index, entity)
        if 0 < level:
            sheet.row_dimensions[index].outline_level = level
        index += 1
    elif has_dhcp_range_or_reserve(entity):
        if etype == Entity.IP4Network or full:
            write_structure_node_for_excel(sheet, level, start_column, index, entity)
            if 0 < level:
                sheet.row_dimensions[index].outline_level = level
            index += 1
            level += 1
        
        nodes = get_sorted_children(entity)
        for node in nodes:
            index = write_structure_for_excel(sheet, level , \
                                              start_column, index, node['entity'], full)
    return index

# --------------------------------------------
# Top Writing function for Excel format.
# --------------------------------------------
def export_as_excel(api, dirname, filename, id, full):
    clear_pool_class_cache()
    excel_config = col_config['excel']
    start_column = excel_config['start_column']
    start_row = excel_config['start_row']

    workbook = openpyxl.load_workbook(dirname + '/templates/Default.xlsx')
    sheet = workbook.active
    entity = api.get_entity_by_id(id)
    sheet.title = get_sheet_name(entity)
    write_structure_title_for_excel(sheet, entity)
    write_structure_header_for_excel(start_column, start_row, sheet)
    add_auto_filters(sheet, start_column, start_row, len(col_config['props']))
    index =  start_row + 1
    if entity.get_type() == Entity.Configuration:
        for ip4_block in entity.get_children_of_type(Entity.IP4Block):
            index = write_structure_for_excel(sheet, 0, \
                                              start_column, index, ip4_block, full)
    else:
        write_structure_for_excel(sheet, 0, start_column, index, entity, full)
    workbook.save(dirname + '/' + filename)
    